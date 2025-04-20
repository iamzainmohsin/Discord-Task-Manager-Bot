import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os


class Task:
    def __init__(self, description, due_date, status, index=0):
        self.index = index
        self.description = description
        self.due_date = self.format_date(due_date)
        self.status = status

    @staticmethod
    def format_date(date_str):
        try:
            # Attempt to parse the date in various common formats
            date_formats = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%B-%Y"]
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime(
                        "%d-%B-%Y"
                    )  # Output in date-Month-Year format
                except ValueError:
                    continue
            raise ValueError("Invalid date format")  # If no format matches
        except Exception as e:
            return f"Error formatting date: {e}"

    # Output!
    def __str__(self):
        # User-friendly representation
        return f"{self.index}.\nTask: '{self.description}'\nDue Date: {self.due_date}\nCurrent status: {self.status}\n"


class ExcelHandler:

    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = Workbook()

    # Set Headers:
    def set_headers(self, worksheet1):
        header_text = ["Task Description", "Due Date", "Status"]
        worksheet1.append(header_text)
        print(f"Created Sheets {worksheet1.title}")

    # sets up the Workbook!
    def workbook_setup(self):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            print(f"Workbook has been loaded!")

            # checks for headers in Already existing sheet:
            if self.workbook:
                workheet = self.workbook["Pending"]

        else:
            del self.workbook["Sheet"]
            worksheet_pending = self.workbook.create_sheet(title="Pending")

            # headers Setup:
            self.set_headers(worksheet_pending)
            self.workbook.save(self.file_name)
            print(f"New Workbook has been created!")

    # Writes to the Workbook!
    def add_tasks(self, task: Task):
        worksheet = self.workbook["Pending"]
        row_data = [task.description, task.due_date, task.status]
        worksheet.append(row_data)
        self.workbook.save(self.file_name)
        print(
            f"DEBUG: Adding task - {task.description}, {task.due_date}, {task.status}"
        )  # Debugging line
        return True

    # Reading from the Workbook!
    def get_tasks(self):
        tasks_list = []
        worksheet = self.workbook["Pending"]

        # Reads the Data:
        for index, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=1
        ):
            try:
                if row and len(row) >= 3:
                    description, due_date, status = row
                    print(f"Task ID: {index}, Description: {description}")
                    task_obj = Task(
                        index=index,
                        description=description,
                        due_date=due_date,
                        status=status,
                    )
                    tasks_list.append(task_obj)
                    print(task_obj)
                else:
                    print(f"Skipping incomplete row: {row}")
            except Exception as e:
                print(f"Error processing row {index}: {e}")
                continue

        return tasks_list

    # Updating Status of the Task
    def complete_task(self, task_id, new_details):
        worksheet = self.workbook["Pending"]
        task_id = int(task_id)

        # Edge Cases fix:
        valid_statuses = ["P", "C"]
        if new_details not in valid_statuses:
            print(
                f"Invalid status '{new_details}'. Use 'C' for Completed or 'P' for Pending."
            )
            return

        for index, row in enumerate(worksheet.iter_rows(min_row=2), start=1):
            if index == task_id:
                row[2].value = new_details  # Row2 == Status column
                self.workbook.save(self.file_name)
                print(f"{task_id} marked as Completed")
                return

    # Moving Completed Tasks
    def delete_completed_task(self):
        pending_worksheet = self.workbook["Pending"]
        rows_to_delete = []

        for row_index, row in enumerate(
            pending_worksheet.iter_rows(min_row=2), start=2
        ):  # Skip header
            if row[2].value == "C":  # Column 3 = Status
                rows_to_delete.append(row_index)

        if not rows_to_delete:
            print("No completed tasks to delete.")
            return

        for row_index in reversed(rows_to_delete):
            pending_worksheet.delete_rows(row_index)
            print(f"Deleted completed task at row {row_index}")

        self.workbook.save(self.file_name)
        print("All completed tasks deleted.")

    # Deleting Tasks:
    def delete_task(self, task_id):
        # Validate the task_id
        if not task_id.isdigit():
            print(f"Invalid Task ID: '{task_id}'. Please choose Numeric Values only")
            return False

        worksheet = self.workbook["Pending"]
        print(f"Searching for Task ID '{task_id}' in sheet: {worksheet}")

        # Start enumerating from 1 while manually skipping the header
        task_deleted = False
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
            if row_index == 1:  # Skip the header row explicitly
                continue

            # Check if task_id matches the row index (adjust for Python's indexing)
            if str(row_index - 1) == str(
                task_id
            ):  # Subtract 1 to align with task numbering logic
                worksheet.delete_rows(row_index)  # Delete the row
                self.workbook.save(self.file_name)  # Save the updated workbook
                print(
                    f"Task ID '{task_id}' successfully deleted from sheet: {worksheet}."
                )
                task_deleted = True
                break

        print(f"Task ID '{task_id}' not found in sheet: {worksheet}.")
        return task_deleted
